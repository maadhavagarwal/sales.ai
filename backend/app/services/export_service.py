"""
Advanced Data Export Service
Provides PDF reports with charts, Excel exports, CSV exports, and Power BI templates
"""

import io
import sqlite3
from datetime import datetime
from typing import Dict, Any, List, Optional, BinaryIO
import pandas as pd
from fpdf import FPDF
import matplotlib.pyplot as plt
import matplotlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Use non-interactive backend for server environments
matplotlib.use('Agg')


class AdvancedPDFReport(FPDF):
    """Enhanced PDF with charts and formatting"""
    
    def __init__(self, title: str = "NeuralBI Report"):
        super().__init__()
        self.title = title
        self.page_title = title
        self.company_name = ""
        
    def header(self):
        """Page header with logo/title"""
        self.set_font("Helvetica", "B", 16)
        self.cell(0, 10, self.page_title, 0, 1, "C")
        if self.company_name:
            self.set_font("Helvetica", "I", 10)
            self.cell(0, 8, f"Company: {self.company_name}", 0, 1, "C")
        self.ln(5)
    
    def footer(self):
        """Page footer with date and page number"""
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.cell(0, 10, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Page {self.page_no()}", 0, 0, "C")
    
    def add_heading(self, text: str, level: int = 1):
        """Add a heading"""
        sizes = {1: 16, 2: 14, 3: 12}
        self.set_font("Helvetica", "B", sizes.get(level, 12))
        self.cell(0, 10, text, 0, 1)
        self.ln(3)
    
    def add_paragraph(self, text: str):
        """Add paragraph with wrapping"""
        self.set_font("Helvetica", "", 11)
        self.multi_cell(0, 6, text)
        self.ln(3)
    
    def add_table(self, df: pd.DataFrame, max_rows: int = 50):
        """Add a table from DataFrame"""
        if len(df) > max_rows:
            df = df.head(max_rows)
        
        self.set_font("Helvetica", "B", 9)
        col_width = 190 / len(df.columns)
        
        # Header
        for col in df.columns:
            col_name = str(col)[:15]  # Truncate long names
            self.cell(col_width, 7, col_name, border=1)
        self.ln()
        
        # Data
        self.set_font("Helvetica", "", 8)
        for idx, row in df.iterrows():
            for val in row:
                cell_val = str(val)[:20]  # Truncate long values
                self.cell(col_width, 6, cell_val, border=1)
            self.ln()
    
    def add_image_from_bytes(self, img_bytes: bytes, width: int = 180):
        """Add image from bytes"""
        try:
            img_io = io.BytesIO(img_bytes)
            img_io.seek(0)
            self.image(img_io, w=width, h=100)
            self.ln(5)
        except Exception as e:
            self.add_paragraph(f"[Chart rendering skipped: {str(e)}]")


class ExportService:
    """Handles all data export formats"""
    
    @staticmethod
    def generate_pdf_report(
        dataset_id: str,
        company_id: str,
        df: pd.DataFrame,
        insights: List[Dict[str, Any]],
        include_charts: bool = True
    ) -> bytes:
        """Generate comprehensive PDF report with charts and data"""
        
        pdf = AdvancedPDFReport(f"Data Analysis Report - {dataset_id}")
        pdf.company_name = company_id
        pdf.alias_nb_pages()
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)
        
        # Executive Summary
        pdf.add_heading("Executive Summary", 1)
        pdf.add_paragraph(f"Dataset: {dataset_id}")
        pdf.add_paragraph(f"Records: {len(df)}")
        pdf.add_paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        pdf.ln(5)
        
        # Key Insights
        if insights:
            pdf.add_heading("Key Insights", 2)
            for idx, insight in enumerate(insights[:10], 1):  # Top 10 insights
                title = insight.get("title", f"Insight {idx}")
                impact = insight.get("impact_score", 0)
                pdf.add_paragraph(f"• {title} (Impact: {impact}%)")
        pdf.ln(5)
        
        # Data Preview
        pdf.add_heading("Data Preview", 2)
        preview_df = df.head(20).copy()
        pdf.add_table(preview_df, max_rows=20)
        pdf.ln(5)
        
        # Statistical Summary
        if len(df.select_dtypes(include=['number']).columns) > 0:
            pdf.add_heading("Statistical Summary", 2)
            stats_df = df.describe().round(2).reset_index()
            pdf.add_table(stats_df)
        pdf.ln(5)
        
        # Charts (if requested and data allows)
        if include_charts and len(df) > 0:
            pdf.add_page()
            pdf.add_heading("Data Visualizations", 2)
            
            # Numeric columns for charting
            numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
            
            if numeric_cols:
                # Create distribution chart
                try:
                    fig, axes = plt.subplots(2, 2, figsize=(10, 8))
                    fig.suptitle('Data Distribution Analysis', fontsize=14, fontweight='bold')
                    
                    for idx, col in enumerate(numeric_cols[:4]):
                        ax = axes[idx // 2, idx % 2]
                        ax.hist(df[col].dropna(), bins=20, edgecolor='black', alpha=0.7)
                        ax.set_title(col)
                        ax.set_xlabel('Value')
                        ax.set_ylabel('Frequency')
                    
                    img_buffer = io.BytesIO()
                    plt.tight_layout()
                    plt.savefig(img_buffer, format='png', dpi=100, bbox_inches='tight')
                    img_buffer.seek(0)
                    plt.close(fig)
                    
                    pdf.add_image_from_bytes(img_buffer.getvalue(), width=180)
                except Exception:
                    pdf.add_paragraph("[Chart generation skipped]")
        
        # Footer note
        pdf.ln(10)
        pdf.set_font("Helvetica", "I", 9)
        pdf.add_paragraph("This report was generated by NeuralBI. For questions, contact support@neuralbi.com")
        
        # Output PDF
        pdf_bytes = pdf.output()
        return pdf_bytes if isinstance(pdf_bytes, bytes) else pdf_bytes.encode('latin-1')
    
    @staticmethod
    def generate_excel_export(
        df: pd.DataFrame,
        filename: str = "export.xlsx",
        include_stats: bool = True,
        include_pivot: bool = False
    ) -> bytes:
        """Generate formatted Excel export with multiple sheets"""
        
        excel_buffer = io.BytesIO()
        
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            # Main data sheet
            df.to_excel(writer, sheet_name='Data', index=False)
            
            # Statistical summary sheet
            if include_stats and len(df.select_dtypes(include=['number']).columns) > 0:
                stats_df = df.describe().T.reset_index()
                stats_df.to_excel(writer, sheet_name='Statistics', index=False)
            
            # Pivot table (if requested)
            if include_pivot:
                numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
                if len(numeric_cols) > 0 and len(df) > 0:
                    try:
                        pivot_df = df.pivot_table(numeric_only=True, aggfunc='mean')
                        pivot_df.to_excel(writer, sheet_name='Pivot')
                    except Exception:
                        pass
            
            # Data Types sheet
            types_df = pd.DataFrame({
                'Column': df.columns,
                'Data Type': [str(dtype) for dtype in df.dtypes]
            })
            types_df.to_excel(writer, sheet_name='Schema', index=False)
        
        # Format workbook
        workbook = writer.book
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            
            # Format headers (first row)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        excel_buffer.seek(0)
        return excel_buffer.getvalue()
    
    @staticmethod
    def generate_csv_export(
        df: pd.DataFrame,
        filename: str = "export.csv",
        include_index: bool = False
    ) -> str:
        """Generate CSV export"""
        return df.to_csv(index=include_index)
    
    @staticmethod
    def generate_power_bi_template(
        dataset_id: str,
        company_id: str,
        df: pd.DataFrame,
        columns_info: Optional[Dict[str, str]] = None
    ) -> Dict[str, Any]:
        """Generate Power BI template configuration JSON"""
        
        numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
        categorical_cols = df.select_dtypes(include=['object']).columns.tolist()
        date_cols = df.select_dtypes(include=['datetime64']).columns.tolist()
        
        template = {
            "name": f"{dataset_id} - Power BI Template",
            "description": f"Auto-generated Power BI template for {dataset_id}",
            "company": company_id,
            "created_at": datetime.now().isoformat(),
            "tables": [
                {
                    "name": "MainData",
                    "description": f"Main data table with {len(df)} records",
                    "columns": [
                        {
                            "name": col,
                            "type": "Numeric" if col in numeric_cols else "Text" if col in categorical_cols else "DateTime",
                            "dataType": str(df[col].dtype),
                            "hidden": False,
                        }
                        for col in df.columns
                    ]
                }
            ],
            "measures": [
                {
                    "name": f"Total {col}",
                    "expression": f"SUM(MainData[{col}])",
                    "dataType": "Decimal",
                    "visible": True,
                }
                for col in numeric_cols[:5]  # Top 5 numeric columns
            ],
            "visualizations": [
                {
                    "title": "Overview",
                    "type": "KPI",
                    "measures": [f"Total {col}" for col in numeric_cols[:3]]
                },
                {
                    "title": "Trends",
                    "type": "LineChart",
                    "axis": date_cols[0] if date_cols else None,
                    "values": numeric_cols[:2]
                },
                {
                    "title": "Distribution",
                    "type": "ColumnChart",
                    "category": categorical_cols[0] if categorical_cols else numeric_cols[0],
                    "values": [numeric_cols[0]] if numeric_cols else []
                }
            ],
            "import_steps": [
                "1. Open Power BI Desktop",
                "2. Create new blank report",
                "3. Import data from CSV or Excel",
                "4. Copy column definitions from this template",
                "5. Create measures listed above",
                "6. Build visualizations as specified",
                "7. Save and publish to Power BI Service"
            ]
        }
        
        return template
    
    @staticmethod
    def generate_json_export(
        df: pd.DataFrame,
        orient: str = "records"  # records, split, index, columns, values
    ) -> str:
        """Generate JSON export"""
        return df.to_json(orient=orient, date_format='iso')


# Export function for main.py integration
def create_dataset_export(
    dataset_id: str,
    company_id: str,
    df: pd.DataFrame,
    insights: Optional[List[Dict[str, Any]]] = None,
    export_format: str = "pdf",
    **kwargs
):
    """
    Main export function supporting multiple formats
    
    Args:
        dataset_id: Name of dataset
        company_id: Company ID for filtering
        df: DataFrame to export
        insights: Optional list of insights to include
        export_format: 'pdf', 'excel', 'csv', 'json', 'power_bi'
        **kwargs: Format-specific options
    
    Returns:
        Exported data (bytes for PDF/Excel, str for CSV/JSON, dict for Power BI)
    """
    
    insights = insights or []
    
    if export_format == "pdf":
        return ExportService.generate_pdf_report(
            dataset_id, company_id, df, insights,
            include_charts=kwargs.get('include_charts', True)
        )
    elif export_format == "excel":
        return ExportService.generate_excel_export(
            df,
            include_stats=kwargs.get('include_stats', True),
            include_pivot=kwargs.get('include_pivot', False)
        )
    elif export_format == "csv":
        return ExportService.generate_csv_export(df)
    elif export_format == "json":
        return ExportService.generate_json_export(
            df,
            orient=kwargs.get('orient', 'records')
        )
    elif export_format == "power_bi":
        return ExportService.generate_power_bi_template(
            dataset_id, company_id, df,
            columns_info=kwargs.get('columns_info')
        )
    else:
        raise ValueError(f"Unsupported export format: {export_format}")
